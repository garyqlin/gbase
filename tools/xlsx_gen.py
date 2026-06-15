# SPDX-License-Identifier: MIT
"""Excel 表格生成工具"""

import json
import logging
import os
import subprocess

from lib.toolkit import tool

logger = logging.getLogger(__name__)


@tool()
async def gen_xlsx(title: str = "表格", sheets: list = None, output_path: str = "") -> dict:
    """
    生成 Excel (.xlsx) 文件。

    Args:
        title: 文件名（不含路径）
        sheets: 工作表列表，每项为 {"name": "Sheet1", "headers": ["列1","列2"], "rows": [["a","b"]]}
                支持 formula: {"type": "formula", "value": "=SUM(A1:A10)"}
        output_path: 输出路径，为空则自动生成

    Returns:
        {"path": "...", "size": 12345}
    """
    if not sheets:
        sheets = [{"name": "Sheet1", "headers": ["A"], "rows": [["示例"]]}]

    if not output_path:
        output_path = os.path.expanduser(f"~/Downloads/{title}.xlsx")

    script = f"""import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

sheets = {json.dumps(sheets)}
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for si, sdata in enumerate(sheets):
    if si == 0:
        ws = wb.active
        ws.title = sdata.get("name", "Sheet1")
    else:
        ws = wb.create_sheet(title=sdata.get("name", f"Sheet{si + 1}"))
    
    headers = sdata.get("headers", [])
    if headers:
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    rows = sdata.get("rows", [])
    for ri, row in enumerate(rows, len(headers) > 0 and 2 or 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, dict) and val.get("type") == "formula":
                cell.value = val["value"]
            else:
                cell.value = val
            cell.border = thin_border
    
    # 列宽自适应
    for col in range(1, max(len(headers), 1) + 1):
        max_len = len(str(headers[col-1])) if headers else 10
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 40)

wb.save({json.dumps(output_path)})
print("OK:" + {json.dumps(output_path)})
"""

    result = subprocess.run(["python3", "-c", script], capture_output=True, text=True, timeout=30)

    if result.returncode != 0:
        return {"error": result.stderr.strip() or result.stdout.strip()}

    if os.path.isfile(output_path):
        size = os.path.getsize(output_path)
        return {"path": output_path, "size": size, "sheets": len(sheets)}
    return {"error": "文件未生成"}
